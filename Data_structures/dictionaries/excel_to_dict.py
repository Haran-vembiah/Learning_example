import openpyxl


def getTestData():
    data = []
    book = openpyxl.load_workbook("test_data.xlsx")
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)
    for i in range(2, sheet.max_row + 1):  # to get rows
        Dict = {}
        for j in range(1, sheet.max_column + 1):  # to get columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        data.append(Dict)
    return data


print(getTestData())
